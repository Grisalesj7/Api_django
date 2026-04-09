from django.shortcuts import render, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.http import HttpResponse
from .models import Cliente, Producto, Pedido, DetallePedido
from django.contrib import messages
from .forms import PedidoForm, DetallePedidoFormSet
from django.db import transaction
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.utils import timezone
import json
from django.db import models


# Librerías para exportar
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

@login_required
def home(request):
    return render(request, 'home.html')

# ---------------------- CLIENTE ----------------------
class ClienteListView(LoginRequiredMixin, ListView):
    model = Cliente
    template_name = 'lista_clientes.html'
    paginate_by = 5

class ClienteCreateView(LoginRequiredMixin, CreateView):
    model = Cliente
    fields = '__all__'
    template_name = 'formulario_general.html'
    success_url = reverse_lazy('cliente_list')

class ClienteUpdateView(LoginRequiredMixin, UpdateView):
    model = Cliente
    fields = '__all__'
    template_name = 'formulario_general.html'
    success_url = reverse_lazy('cliente_list')

class ClienteDeleteView(LoginRequiredMixin, DeleteView):
    model = Cliente
    success_url = reverse_lazy('cliente_list')

# --- FUNCIONES DE EXPORTACIÓN INDIVIDUAL ---
@login_required
def exportar_cliente_excel(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Datos_{cliente.nombre}"

    # Encabezados
    ws.append(['ID', 'Nombre', 'Correo', 'Dirección', 'Teléfono'])
    # Datos
    ws.append([cliente.id, cliente.nombre, cliente.correo, cliente.direccion, cliente.telefono])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Cliente_{cliente.nombre}.xlsx'
    wb.save(response)
    return response

@login_required
def exportar_cliente_pdf(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Ficha_{cliente.nombre}.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    
    # Título
    p.setFont("Helvetica-Bold", 18)
    p.drawString(100, 750, f"Ficha de Cliente: {cliente.nombre}")
    
    # Cuerpo
    p.setFont("Helvetica", 12)
    p.drawString(100, 710, f"ID: {cliente.id}")
    p.drawString(100, 690, f"Correo electrónico: {cliente.correo}")
    p.drawString(100, 670, f"Dirección: {cliente.direccion}")
    p.drawString(100, 650, f"Teléfono: {cliente.telefono}")
    
    p.line(100, 630, 500, 630) # Línea divisoria
    
    # Pie de página
    p.setFont("Helvetica-Oblique", 9)
    p.drawString(100, 50, "Sistema de Gestión de Pedidos - Reporte Individual")

    p.showPage()
    p.save()
    return response

# ---------------------- RESTO DE VISTAS (Producto, Pedido) ----------------------
class ProductoListView(LoginRequiredMixin, ListView):
    model = Producto
    template_name = 'lista_productos.html'
    paginate_by = 5

class ProductoCreateView(LoginRequiredMixin, CreateView):
    model = Producto
    fields = '__all__'
    template_name = 'formulario_general.html'
    success_url = reverse_lazy('producto_list')

class ProductoUpdateView(LoginRequiredMixin, UpdateView):
    model = Producto
    fields = '__all__'
    template_name = 'formulario_general.html'
    success_url = reverse_lazy('producto_list')

class ProductoDeleteView(LoginRequiredMixin, DeleteView):
    model = Producto
    success_url = reverse_lazy('producto_list')

class PedidoListView(LoginRequiredMixin, ListView):
    model = Pedido
    template_name = 'lista_pedidos.html'
    paginate_by = 5

class PedidoCreateView(LoginRequiredMixin, CreateView):
    model = Pedido
    fields = '__all__'
    template_name = 'formulario_general.html'
    success_url = reverse_lazy('pedido_list')

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data['detalles'] = DetallePedidoFormSet(self.request.POST)
        else:
            data['detalles'] = DetallePedidoFormSet()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        detalles = context['detalles']
        
        # Usamos una transacción para que si algo falla en los productos, 
        # no se guarde el pedido a medias.
        with transaction.atomic():
            self.object = form.save()
            if detalles.is_valid():
                detalles.instance = self.object
                detalles.save()
                
                # Opcional: Calcular el total_pagado automáticamente al guardar
                total = sum(item.subtotal for item in self.object.detalles.all())
                self.object.total_pagado = total
                self.object.save()
            else:
                return self.form_invalid(form)
                
        return super().form_valid(form)

class PedidoUpdateView(LoginRequiredMixin, UpdateView):
    model = Pedido
    fields = '__all__'
    template_name = 'formulario_general.html'
    success_url = reverse_lazy('pedido_list')

class PedidoDeleteView(LoginRequiredMixin, DeleteView):
    model = Pedido
    success_url = reverse_lazy('pedido_list')

    def post(self, request, *args, **kwargs):
        # Obtenemos el pedido que se intenta eliminar
        self.object = self.get_object()
        
        # Validación: Si el estado es Enviado o Entregado, bloqueamos el borrado
        # Asegúrate de que "Enviado" y "Entregado" coincidan exactamente con lo que tienes en tu Base de Datos
        if self.object.estado in ['Enviado', 'Entregado']:
            messages.error(request, f"No se puede eliminar el pedido #{self.object.id} porque ya está {self.object.estado}.")
            return render(request, 'pedidos/pedido_confirm_delete.html', {'object': self.object})
        
        # Si no está en esos estados, procedemos a borrar
        return super().post(request, *args, **kwargs)

# --- EXCEL DE PRODUCTO INDIVIDUAL ---
@login_required
def exportar_producto_excel(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Producto_{producto.id}"

    # Cabeceras
    ws.append(['ID', 'Nombre del Producto', 'Precio', 'Stock Actual'])
    # Datos
    ws.append([producto.id, producto.nombre, f"${producto.precio}", producto.stock])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Producto_{producto.nombre}.xlsx'
    wb.save(response)
    return response

# --- PDF DE PRODUCTO INDIVIDUAL ---
@login_required
def exportar_producto_pdf(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Ficha_Producto_{producto.nombre}.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    
    # Encabezado estilo ficha técnica
    p.setFont("Helvetica-Bold", 18)
    p.drawString(100, 750, "FICHA TÉCNICA DE PRODUCTO")
    
    p.setStrokeColorRGB(0.2, 0.2, 0.2)
    p.line(100, 740, 500, 740)

    p.setFont("Helvetica", 12)
    p.drawString(100, 710, f"Nombre: {producto.nombre}")
    p.drawString(100, 690, f"Precio Unitario: ${producto.precio}")
    p.drawString(100, 670, f"Existencias en Inventario: {producto.stock} unidades")
    
    # Estado del stock
    if producto.stock <= 5:
        p.setFillColorRGB(0.8, 0, 0) # Rojo si hay poco stock
        p.drawString(100, 640, "ESTADO: ¡STOCK CRÍTICO!")
    else:
        p.setFillColorRGB(0, 0.5, 0) # Verde si está bien
        p.drawString(100, 640, "ESTADO: Stock Disponible")

    p.showPage()
    p.save()
    return response

# --- EXCEL DE PEDIDO INDIVIDUAL ---
@login_required
def exportar_pedido_pdf(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    detalles = pedido.detalles.all()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Factura_Pedido_{pedido.id}.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    
    # --- ENCABEZADO ESTILO PREMIUM ---
    p.setFillColorRGB(0.96, 0.91, 0.84) # Beige suave (Estilo wellness)
    p.rect(0, 740, 612, 50, fill=1, stroke=0)
    
    p.setFillColorRGB(0.2, 0.2, 0.2) # Gris oscuro para el texto
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, 760, f"ORDEN DE PEDIDO #{pedido.id}")
    
    # --- DATOS DEL CLIENTE ---
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, 710, "CLIENTE:")
    p.setFont("Helvetica", 11)
    p.drawString(110, 710, f"{pedido.cliente.nombre}")
    p.drawString(50, 690, f"FECHA: {pedido.fecha.strftime('%d/%m/%Y')}")
    p.drawString(50, 675, f"ESTADO: {pedido.estado.upper()}")
    
    # --- TABLA DE PRODUCTOS ---
    y = 630
    p.setFont("Helvetica-Bold", 11)
    p.drawString(50, y, "PRODUCTO")
    p.drawString(350, y, "CANT.")
    p.drawString(450, y, "SUBTOTAL")
    
    p.setStrokeColorRGB(0.8, 0.8, 0.8)
    p.line(50, y-5, 550, y-5)
    
    y -= 25
    p.setFont("Helvetica", 11)
    for item in detalles:
        p.drawString(50, y, f"{item.producto.nombre[:40]}")
        p.drawString(350, y, f"{item.cantidad}")
        p.drawString(450, y, f"${item.subtotal:,.0f}")
        y -= 20
        
        if y < 100: # Control de página
            p.showPage()
            y = 750
    
    # --- RESUMEN FINAL ---
    y -= 20
    p.setStrokeColorRGB(0.2, 0.2, 0.2)
    p.line(350, y+15, 550, y+15)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(350, y, "TOTAL A PAGAR:")
    p.drawString(480, y, f"${(pedido.total_pagado or 0):,.0f}")

    p.showPage()
    p.save()
    return response

# --- EXCEL DE PEDIDO INDIVIDUAL ---
@login_required
def exportar_pedido_excel(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    detalles = pedido.detalles.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pedido_{pedido.id}"

    # Encabezados
    ws.append(['ID Pedido', 'Cliente', 'Fecha', 'Estado', 'Total'])
    ws.append([pedido.id, pedido.cliente.nombre, pedido.fecha, pedido.estado, pedido.total_pagado])
    
    ws.append([]) # Espacio en blanco
    
    # Detalles del pedido
    ws.append(['Producto', 'Cantidad', 'Precio Unitario', 'Subtotal'])
    for item in detalles:
        ws.append([item.producto.nombre, item.cantidad, item.producto.precio, item.subtotal])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Pedido_{pedido.id}.xlsx'
    wb.save(response)
    return response

@login_required
def dashboard(request):
    # 1. Configuración de fechas (Mes actual)
    ahora = timezone.now()
    primer_dia_mes = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # 2. CÁLCULO: Top 5 Productos más vendidos del mes
    # Filtramos los detalles de pedidos que pertenecen a este mes, 
    # agrupamos por nombre de producto y sumamos sus cantidades.
    productos_top_raw = DetallePedido.objects.filter(
        pedido__fecha__gte=primer_dia_mes
    ).values(
        'producto__nombre'
    ).annotate(
        total_vendido=Sum('cantidad')
    ).order_by('-total_vendido')[:5]

    # Preparamos las listas para la gráfica (Chart.js necesita JSON)
    labels_grafica = [item['producto__nombre'] for item in productos_top_raw]
    datos_grafica = [item['total_vendido'] for item in productos_top_raw]

    # 3. CÁLCULO: Resumen de Ventas (Dinero)
    ventas_mes = Pedido.objects.filter(
        fecha__gte=primer_dia_mes
    ).aggregate(total=Sum('total_pagado'))['total'] or 0

    # 4. CÁLCULO: Stock Crítico
    # Productos que tienen 5 o menos unidades
    stock_critico_count = Producto.objects.filter(stock__lte=5).count()

    context = {
        # Widgets de conteo rápido
        'clientes_total': Cliente.objects.count(),
        'productos_total': Producto.objects.count(),
        'pedidos_total': Pedido.objects.count(),
        'ventas_mes_monto': ventas_mes,
        'stock_critico': stock_critico_count,
        
        # Datos para las gráficas
        'grafica_labels': json.dumps(labels_grafica),
        'grafica_datos': json.dumps(datos_grafica),
        
        # Lista de actividad reciente
        'ultimos': Pedido.objects.order_by('-id')[:5],
        'mes_nombre': ahora.strftime('%B')
    }
    return render(request, 'dashboard.html', context)



# ====================== API REST (DRF) ======================

from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import IsAuthenticated
from .serializers import ClienteSerializer, ProductoSerializer, PedidoSerializer, DetallePedidoSerializer


# 🔵 API CLIENTES
class ClienteAPI(ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer
    permission_classes = [IsAuthenticated]


# 🔵 API PRODUCTOS
class ProductoAPI(ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated]


# 🔵 API PEDIDOS
class PedidoAPI(ModelViewSet):
    queryset = Pedido.objects.all()
    serializer_class = PedidoSerializer
    permission_classes = [IsAuthenticated]

# 🔵 API DETALLES DE PEDIDO
class DetallePedidoAPI(ModelViewSet):
    queryset = DetallePedido.objects.all()
    serializer_class = DetallePedidoSerializer
    permission_classes = [IsAuthenticated]